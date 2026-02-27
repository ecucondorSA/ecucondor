"""
ECUCONDOR - Exportadores PDF y Excel
Genera reportes contables en formatos PDF y Excel.
"""

from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer


class ExportadorPDF:
    """Genera reportes contables en formato PDF."""

    def __init__(self, empresa: str = "ECUCONDOR", ruc: str = ""):
        self.empresa = empresa
        self.ruc = ruc
        self.styles = getSampleStyleSheet()

    def _crear_estilo_titulo(self) -> ParagraphStyle:
        return ParagraphStyle(
            'Titulo',
            parent=self.styles['Heading1'],
            fontSize=14,
            alignment=1,  # Center
            spaceAfter=12
        )

    def _crear_encabezado(self, titulo: str, fecha: str) -> list:
        """Crea encabezado comun para todos los reportes."""
        estilo_titulo = self._crear_estilo_titulo()
        estilo_empresa = ParagraphStyle(
            'Empresa',
            parent=self.styles['Normal'],
            fontSize=12,
            alignment=1,
            spaceAfter=6
        )
        estilo_info = ParagraphStyle(
            'Info',
            parent=self.styles['Normal'],
            fontSize=10,
            alignment=1,
            spaceAfter=20
        )

        return [
            Paragraph(self.empresa, estilo_empresa),
            Paragraph(f"RUC: {self.ruc}", estilo_info) if self.ruc else Spacer(1, 0),
            Paragraph(titulo, estilo_titulo),
            Paragraph(f"Fecha: {fecha}", estilo_info),
            Spacer(1, 12),
        ]

    def exportar_balance(self, balance: Any, fecha_corte: date) -> bytes:
        """Exporta Balance General a PDF."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
        elementos = []

        # Encabezado
        elementos.extend(self._crear_encabezado(
            "BALANCE GENERAL",
            fecha_corte.strftime("%d/%m/%Y")
        ))

        # Tabla de Activos
        datos_activos = [["ACTIVOS", "", ""]]
        for linea in balance.activos.lineas:
            if linea.es_titulo:
                datos_activos.append([linea.nombre.upper(), "", ""])
            else:
                datos_activos.append([f"  {linea.codigo} - {linea.nombre}", "", f"${linea.saldo:,.2f}"])

        datos_activos.append(["TOTAL ACTIVOS", "", f"${float(balance.total_activos):,.2f}"])

        # Tabla de Pasivos
        datos_pasivos = [["PASIVOS", "", ""]]
        for linea in balance.pasivos.lineas:
            if linea.es_titulo:
                datos_pasivos.append([linea.nombre.upper(), "", ""])
            else:
                datos_pasivos.append([f"  {linea.codigo} - {linea.nombre}", "", f"${linea.saldo:,.2f}"])

        datos_pasivos.append(["TOTAL PASIVOS", "", f"${float(balance.total_pasivos):,.2f}"])

        # Tabla de Patrimonio
        datos_patrimonio = [["PATRIMONIO", "", ""]]
        for linea in balance.patrimonio.lineas:
            if linea.es_titulo:
                datos_patrimonio.append([linea.nombre.upper(), "", ""])
            else:
                datos_patrimonio.append([f"  {linea.codigo} - {linea.nombre}", "", f"${linea.saldo:,.2f}"])

        datos_patrimonio.append(["TOTAL PATRIMONIO", "", f"${float(balance.total_patrimonio):,.2f}"])
        datos_patrimonio.append(["", "", ""])
        datos_patrimonio.append(["TOTAL PASIVO + PATRIMONIO", "", f"${float(balance.total_pasivos + balance.total_patrimonio):,.2f}"])

        # Estilo de tablas
        estilo_tabla = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d3b66')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
        ])

        # Crear tablas
        tabla_activos = Table(datos_activos, colWidths=[10*cm, 2*cm, 4*cm])
        tabla_activos.setStyle(estilo_tabla)
        elementos.append(tabla_activos)
        elementos.append(Spacer(1, 12))

        tabla_pasivos = Table(datos_pasivos, colWidths=[10*cm, 2*cm, 4*cm])
        tabla_pasivos.setStyle(estilo_tabla)
        elementos.append(tabla_pasivos)
        elementos.append(Spacer(1, 12))

        tabla_patrimonio = Table(datos_patrimonio, colWidths=[10*cm, 2*cm, 4*cm])
        tabla_patrimonio.setStyle(estilo_tabla)
        elementos.append(tabla_patrimonio)

        doc.build(elementos)
        return buffer.getvalue()

    def exportar_resultados(self, estado: Any, fecha_inicio: date, fecha_fin: date) -> bytes:
        """Exporta Estado de Resultados a PDF."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
        elementos = []

        # Encabezado
        elementos.extend(self._crear_encabezado(
            "ESTADO DE RESULTADOS",
            f"{fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}"
        ))

        # Tabla de Ingresos
        datos = [["CONCEPTO", "VALOR"]]

        datos.append(["INGRESOS", ""])
        for linea in estado.ingresos:
            if linea.es_titulo:
                datos.append([linea.nombre.upper(), ""])
            else:
                datos.append([f"  {linea.codigo} - {linea.nombre}", f"${linea.saldo:,.2f}"])

        datos.append(["TOTAL INGRESOS", f"${float(estado.total_ingresos):,.2f}"])
        datos.append(["", ""])

        datos.append(["GASTOS", ""])
        for linea in estado.gastos:
            if linea.es_titulo:
                datos.append([linea.nombre.upper(), ""])
            else:
                datos.append([f"  {linea.codigo} - {linea.nombre}", f"${linea.saldo:,.2f}"])

        datos.append(["TOTAL GASTOS", f"${float(estado.total_gastos):,.2f}"])
        datos.append(["", ""])

        resultado = "UTILIDAD" if estado.es_utilidad else "PERDIDA"
        datos.append([f"{resultado} DEL EJERCICIO", f"${float(estado.utilidad_bruta):,.2f}"])

        # Estilo
        estilo_tabla = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d3b66')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
        ])

        tabla = Table(datos, colWidths=[12*cm, 4*cm])
        tabla.setStyle(estilo_tabla)
        elementos.append(tabla)

        doc.build(elementos)
        return buffer.getvalue()

    def exportar_libro_diario(self, asientos: list, periodo: str) -> bytes:
        """Exporta Libro Diario a PDF."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
        elementos = []

        # Encabezado
        elementos.extend(self._crear_encabezado(
            "LIBRO DIARIO",
            periodo
        ))

        # Tabla
        datos = [["FECHA", "NO. ASIENTO", "CONCEPTO", "DEBE", "HABER"]]

        for asiento in asientos:
            datos.append([
                asiento.get('fecha', ''),
                asiento.get('numero_asiento', ''),
                asiento.get('concepto', '')[:50],
                f"${asiento.get('total_debe', 0):,.2f}",
                f"${asiento.get('total_haber', 0):,.2f}",
            ])

        estilo_tabla = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d3b66')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (3, 0), (4, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ])

        tabla = Table(datos, colWidths=[2.5*cm, 2.5*cm, 7*cm, 2.5*cm, 2.5*cm])
        tabla.setStyle(estilo_tabla)
        elementos.append(tabla)

        doc.build(elementos)
        return buffer.getvalue()


class ExportadorExcel:
    """Genera reportes contables en formato Excel."""

    def __init__(self, empresa: str = "ECUCONDOR", ruc: str = ""):
        self.empresa = empresa
        self.ruc = ruc

    def _aplicar_estilo_encabezado(self, ws, fila: int, cols: int):
        """Aplica estilo al encabezado."""
        fill = PatternFill(start_color="0D3B66", end_color="0D3B66", fill_type="solid")
        font = Font(color="FFFFFF", bold=True)
        for col in range(1, cols + 1):
            cell = ws.cell(row=fila, column=col)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal='center')

    def _ajustar_anchos(self, ws):
        """Ajusta ancho de columnas automaticamente."""
        for column_cells in ws.columns:
            length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)

    def exportar_balance(self, balance: Any, fecha_corte: date) -> bytes:
        """Exporta Balance General a Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Balance General"

        # Encabezado
        ws['A1'] = self.empresa
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"RUC: {self.ruc}"
        ws['A3'] = "BALANCE GENERAL"
        ws['A3'].font = Font(bold=True, size=12)
        ws['A4'] = f"Al {fecha_corte.strftime('%d/%m/%Y')}"

        fila = 6

        # Activos
        ws.cell(row=fila, column=1, value="ACTIVOS")
        ws.cell(row=fila, column=1).font = Font(bold=True)
        self._aplicar_estilo_encabezado(ws, fila, 3)
        fila += 1

        for linea in balance.activos.lineas:
            if linea.es_titulo:
                ws.cell(row=fila, column=1, value=linea.nombre.upper()).font = Font(bold=True)
            else:
                ws.cell(row=fila, column=1, value=f"  {linea.codigo}")
                ws.cell(row=fila, column=2, value=linea.nombre)
                ws.cell(row=fila, column=3, value=float(linea.saldo))
                ws.cell(row=fila, column=3).number_format = '#,##0.00'
            fila += 1

        ws.cell(row=fila, column=1, value="TOTAL ACTIVOS").font = Font(bold=True)
        ws.cell(row=fila, column=3, value=float(balance.total_activos)).font = Font(bold=True)
        ws.cell(row=fila, column=3).number_format = '#,##0.00'
        fila += 2

        # Pasivos
        ws.cell(row=fila, column=1, value="PASIVOS")
        ws.cell(row=fila, column=1).font = Font(bold=True)
        self._aplicar_estilo_encabezado(ws, fila, 3)
        fila += 1

        for linea in balance.pasivos.lineas:
            if linea.es_titulo:
                ws.cell(row=fila, column=1, value=linea.nombre.upper()).font = Font(bold=True)
            else:
                ws.cell(row=fila, column=1, value=f"  {linea.codigo}")
                ws.cell(row=fila, column=2, value=linea.nombre)
                ws.cell(row=fila, column=3, value=float(linea.saldo))
                ws.cell(row=fila, column=3).number_format = '#,##0.00'
            fila += 1

        ws.cell(row=fila, column=1, value="TOTAL PASIVOS").font = Font(bold=True)
        ws.cell(row=fila, column=3, value=float(balance.total_pasivos)).font = Font(bold=True)
        ws.cell(row=fila, column=3).number_format = '#,##0.00'
        fila += 2

        # Patrimonio
        ws.cell(row=fila, column=1, value="PATRIMONIO")
        ws.cell(row=fila, column=1).font = Font(bold=True)
        self._aplicar_estilo_encabezado(ws, fila, 3)
        fila += 1

        for linea in balance.patrimonio.lineas:
            if linea.es_titulo:
                ws.cell(row=fila, column=1, value=linea.nombre.upper()).font = Font(bold=True)
            else:
                ws.cell(row=fila, column=1, value=f"  {linea.codigo}")
                ws.cell(row=fila, column=2, value=linea.nombre)
                ws.cell(row=fila, column=3, value=float(linea.saldo))
                ws.cell(row=fila, column=3).number_format = '#,##0.00'
            fila += 1

        ws.cell(row=fila, column=1, value="TOTAL PATRIMONIO").font = Font(bold=True)
        ws.cell(row=fila, column=3, value=float(balance.total_patrimonio)).font = Font(bold=True)
        ws.cell(row=fila, column=3).number_format = '#,##0.00'

        self._ajustar_anchos(ws)

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def exportar_resultados(self, estado: Any, fecha_inicio: date, fecha_fin: date) -> bytes:
        """Exporta Estado de Resultados a Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Estado de Resultados"

        # Encabezado
        ws['A1'] = self.empresa
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"RUC: {self.ruc}"
        ws['A3'] = "ESTADO DE RESULTADOS"
        ws['A3'].font = Font(bold=True, size=12)
        ws['A4'] = f"Del {fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}"

        fila = 6

        # Encabezado tabla
        ws.cell(row=fila, column=1, value="CONCEPTO")
        ws.cell(row=fila, column=2, value="VALOR")
        self._aplicar_estilo_encabezado(ws, fila, 2)
        fila += 1

        # Ingresos
        ws.cell(row=fila, column=1, value="INGRESOS").font = Font(bold=True)
        fila += 1

        for linea in estado.ingresos:
            if linea.es_titulo:
                ws.cell(row=fila, column=1, value=linea.nombre.upper()).font = Font(bold=True)
            else:
                ws.cell(row=fila, column=1, value=f"  {linea.codigo} - {linea.nombre}")
                ws.cell(row=fila, column=2, value=float(linea.saldo))
                ws.cell(row=fila, column=2).number_format = '#,##0.00'
            fila += 1

        ws.cell(row=fila, column=1, value="TOTAL INGRESOS").font = Font(bold=True)
        ws.cell(row=fila, column=2, value=float(estado.total_ingresos)).font = Font(bold=True)
        ws.cell(row=fila, column=2).number_format = '#,##0.00'
        fila += 2

        # Gastos
        ws.cell(row=fila, column=1, value="GASTOS").font = Font(bold=True)
        fila += 1

        for linea in estado.gastos:
            if linea.es_titulo:
                ws.cell(row=fila, column=1, value=linea.nombre.upper()).font = Font(bold=True)
            else:
                ws.cell(row=fila, column=1, value=f"  {linea.codigo} - {linea.nombre}")
                ws.cell(row=fila, column=2, value=float(linea.saldo))
                ws.cell(row=fila, column=2).number_format = '#,##0.00'
            fila += 1

        ws.cell(row=fila, column=1, value="TOTAL GASTOS").font = Font(bold=True)
        ws.cell(row=fila, column=2, value=float(estado.total_gastos)).font = Font(bold=True)
        ws.cell(row=fila, column=2).number_format = '#,##0.00'
        fila += 2

        # Resultado
        resultado = "UTILIDAD DEL EJERCICIO" if estado.es_utilidad else "PERDIDA DEL EJERCICIO"
        ws.cell(row=fila, column=1, value=resultado).font = Font(bold=True, size=12)
        ws.cell(row=fila, column=2, value=float(estado.utilidad_bruta)).font = Font(bold=True, size=12)
        ws.cell(row=fila, column=2).number_format = '#,##0.00'

        self._ajustar_anchos(ws)

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def exportar_libro_diario(self, asientos: list, periodo: str) -> bytes:
        """Exporta Libro Diario a Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Libro Diario"

        # Encabezado
        ws['A1'] = self.empresa
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"RUC: {self.ruc}"
        ws['A3'] = "LIBRO DIARIO"
        ws['A3'].font = Font(bold=True, size=12)
        ws['A4'] = f"Periodo: {periodo}"

        fila = 6

        # Encabezado tabla
        headers = ["FECHA", "NO. ASIENTO", "CONCEPTO", "DEBE", "HABER"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=fila, column=col, value=header)
        self._aplicar_estilo_encabezado(ws, fila, len(headers))
        fila += 1

        total_debe = 0
        total_haber = 0

        for asiento in asientos:
            ws.cell(row=fila, column=1, value=asiento.get('fecha', ''))
            ws.cell(row=fila, column=2, value=asiento.get('numero_asiento', ''))
            ws.cell(row=fila, column=3, value=asiento.get('concepto', ''))
            debe = float(asiento.get('total_debe', 0))
            haber = float(asiento.get('total_haber', 0))
            ws.cell(row=fila, column=4, value=debe)
            ws.cell(row=fila, column=4).number_format = '#,##0.00'
            ws.cell(row=fila, column=5, value=haber)
            ws.cell(row=fila, column=5).number_format = '#,##0.00'
            total_debe += debe
            total_haber += haber
            fila += 1

        # Totales
        ws.cell(row=fila, column=3, value="TOTALES").font = Font(bold=True)
        ws.cell(row=fila, column=4, value=total_debe).font = Font(bold=True)
        ws.cell(row=fila, column=4).number_format = '#,##0.00'
        ws.cell(row=fila, column=5, value=total_haber).font = Font(bold=True)
        ws.cell(row=fila, column=5).number_format = '#,##0.00'

        self._ajustar_anchos(ws)

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def exportar_libro_mayor(self, libro: Any, periodo: str) -> bytes:
        """Exporta Libro Mayor a Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Libro Mayor"

        # Encabezado
        ws['A1'] = self.empresa
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"RUC: {self.ruc}"
        ws['A3'] = "LIBRO MAYOR"
        ws['A3'].font = Font(bold=True, size=12)
        ws['A4'] = f"Cuenta: {libro.cuenta_codigo} - {libro.cuenta_nombre}"
        ws['A5'] = f"Periodo: {periodo}"

        fila = 7

        # Encabezado tabla
        headers = ["FECHA", "NO. ASIENTO", "CONCEPTO", "DEBE", "HABER", "SALDO"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=fila, column=col, value=header)
        self._aplicar_estilo_encabezado(ws, fila, len(headers))
        fila += 1

        for mov in libro.movimientos:
            ws.cell(row=fila, column=1, value=str(mov.fecha))
            ws.cell(row=fila, column=2, value=mov.numero_asiento)
            ws.cell(row=fila, column=3, value=mov.concepto)
            ws.cell(row=fila, column=4, value=float(mov.debe))
            ws.cell(row=fila, column=4).number_format = '#,##0.00'
            ws.cell(row=fila, column=5, value=float(mov.haber))
            ws.cell(row=fila, column=5).number_format = '#,##0.00'
            ws.cell(row=fila, column=6, value=float(mov.saldo))
            ws.cell(row=fila, column=6).number_format = '#,##0.00'
            fila += 1

        # Totales
        ws.cell(row=fila, column=3, value="SALDO FINAL").font = Font(bold=True)
        ws.cell(row=fila, column=4, value=float(libro.total_debe)).font = Font(bold=True)
        ws.cell(row=fila, column=4).number_format = '#,##0.00'
        ws.cell(row=fila, column=5, value=float(libro.total_haber)).font = Font(bold=True)
        ws.cell(row=fila, column=5).number_format = '#,##0.00'
        ws.cell(row=fila, column=6, value=float(libro.saldo_final)).font = Font(bold=True)
        ws.cell(row=fila, column=6).number_format = '#,##0.00'

        self._ajustar_anchos(ws)

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
